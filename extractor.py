"""
Excel Extractor
---------------
Reads an .xlsx / .xls file into a normalized 2D grid,
handling merged cells and propagating their values.
Also tracks row-span per cell so the engine can set duration correctly.
"""

from pathlib import Path
from typing import List

import openpyxl
import pandas as pd

from utils import safe_str


class ExcelExtractor:
    """Reads Excel files into a clean 2D string grid."""

    def extract(self, filepath: Path):
        """
        Returns a list of sheets.
        Each sheet is a dict with:
          'grid'     : 2D list of strings (merged cells expanded)
          'row_span' : 2D list of ints — how many rows each cell spans (1 = normal)
        """
        suffix = filepath.suffix.lower()
        if suffix == ".xlsx":
            return self._extract_xlsx(filepath)
        elif suffix == ".xls":
            return self._extract_xls(filepath)
        raise ValueError(f"Unsupported file type: {suffix}")

    # ---------- XLSX ----------

    def _extract_xlsx(self, filepath: Path):
        wb = openpyxl.load_workbook(filepath, data_only=True)
        return [self._sheet_to_grid(ws) for ws in wb.worksheets]

    def _sheet_to_grid(self, ws):
        max_row, max_col = ws.max_row, ws.max_column
        grid = [[safe_str(ws.cell(r, c).value) for c in range(1, max_col + 1)]
                for r in range(1, max_row + 1)]
        # row_span[r][c] = number of rows this cell spans (only set on top-left of merge)
        row_span = [[1] * max_col for _ in range(max_row)]

        for merged_range in ws.merged_cells.ranges:
            min_col, min_row = merged_range.min_col, merged_range.min_row
            max_c, max_r = merged_range.max_col, merged_range.max_row
            anchor = safe_str(ws.cell(min_row, min_col).value)
            span = max_r - min_row + 1
            for r in range(min_row, max_r + 1):
                for c in range(min_col, max_c + 1):
                    grid[r - 1][c - 1] = anchor
            # Record span only on the anchor cell
            row_span[min_row - 1][min_col - 1] = span

        return {"grid": grid, "row_span": row_span}

    # ---------- XLS ----------

    def _extract_xls(self, filepath: Path):
        xls = pd.ExcelFile(filepath)
        sheets = []
        for name in xls.sheet_names:
            df = xls.parse(name, header=None, dtype=str).fillna("")
            grid = [[safe_str(v) for v in row] for row in df.values.tolist()]
            row_span = [[1] * len(grid[0]) for _ in grid] if grid else []
            sheets.append({"grid": grid, "row_span": row_span})
        return sheets